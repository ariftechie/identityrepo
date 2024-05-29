####################################################################
#Script to read and update the OAuth client ID using Ping Admin API#
####################################################################

import requests
import json
import string
import logging
import openpyxl

#ping_base_url
base_url = "https://pingfederate-admin-api.dev-corp.us1.ping.cloud"


#Setting up excel configuration
path = "C:/Repo/Documents/sample.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
row = sheet_obj.max_row
column = sheet_obj.max_column

print("Total Rows: ", row)
print("Total Columns:", column)

#Closing the opened excel
f = open("C:/Repo/Documents/sample.csv", "r")
print(f.read())
f.close

#Creating Logger
logging.basicConfig(filename="C:/Repo/Documents/sample.csv", format='%(asctime)s' - '%(name)s' - '%(levelname)s' - '%(message)s', filemode='w')
logger = logging.getLogger().setLevel("INFO")


#GET Request
def get_request(client_id):
    url = base_url + "/pf-admin-api/v1/oauth/clients/" + client_id
    response = requests.get(url, headers={"X-Xsrf-Header":"PingFederate"}, auth=("arif.technerd@gmail.com", "Password@123"))
    if response.status_code == 404:
        logging.info('client not found: ' + client_id)

    else:
        #assert response.status_code = 200
        json_data = response.json()
        json_str = json.dumps(json_data, indent=4)
        with open("C:/Repo/Documents/sample.csv","w") as f:
            f.write(json_str)
        logging.info('Read client id: ' + client_id)

#call method - reading client from excel
for i in range(1, row + 1):
    cell_obj = sheet_obj.cell(row=i, column=i)
    get_request(cell_obj.value)